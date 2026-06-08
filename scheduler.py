"""
FraudShield — Scheduler Tự động (Zero-touch) v2
================================================
Chạy: python scheduler.py

Luồng:
  Cứ mỗi SCAN_INTERVAL_SECONDS giây:
    1. Scan thư mục inbox/ tìm file CSV mới
    2. Nếu có → chạy pipeline: cleaning → features → predict
    3. Phân loại risk: LOW / MEDIUM / HIGH
    4. Xuất báo cáo HTML vào output/reports/
    5. Xuất báo cáo Excel vào output/
    6. Di chuyển file đã xử lý vào inbox/processed/
    7. Ghi log vào output/scheduler.log
"""

import os
import sys
import time
import shutil
import json
import io
import numpy as np
import pandas as pd
import joblib
import importlib.util
import logging
from datetime import datetime
from pathlib import Path

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
SRC_DIR   = os.path.join(BASE_DIR, "src")
sys.path.insert(0, BASE_DIR)
sys.path.insert(0, SRC_DIR)

# Load config
try:
    from config import (
        SCAN_INTERVAL_SECONDS, INBOX_DIR, ARCHIVE_DIR, LOG_FILE,
        HIGH_RISK_THRESHOLD, MED_RISK_THRESHOLD, MIN_HIGH_TO_ALERT,
    )
except ImportError:
    SCAN_INTERVAL_SECONDS = 60
    INBOX_DIR   = "inbox"
    ARCHIVE_DIR = "inbox/processed"
    LOG_FILE    = "output/scheduler.log"
    HIGH_RISK_THRESHOLD = 0.60
    MED_RISK_THRESHOLD  = 0.30
    MIN_HIGH_TO_ALERT   = 1

INBOX_PATH   = os.path.join(BASE_DIR, INBOX_DIR)
ARCHIVE_PATH = os.path.join(BASE_DIR, ARCHIVE_DIR)
LOG_PATH     = os.path.join(BASE_DIR, LOG_FILE)
MODEL_PATH   = os.path.join(BASE_DIR, "model", "fraud_model.pkl")
META_PATH    = os.path.join(BASE_DIR, "model", "model_meta.json")
OUTPUT_DIR   = os.path.join(BASE_DIR, "output")
REPORTS_DIR  = os.path.join(BASE_DIR, "output", "reports")  # ← Fix: dùng absolute path

os.makedirs(OUTPUT_DIR,   exist_ok=True)
os.makedirs(INBOX_PATH,   exist_ok=True)
os.makedirs(ARCHIVE_PATH, exist_ok=True)
os.makedirs(REPORTS_DIR,  exist_ok=True)  # ← Tạo sẵn thư mục reports

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
    ],
)
log = logging.getLogger("FraudShield")


def _load_module(name, path):
    spec = importlib.util.spec_from_file_location(name, path)
    mod  = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod

_cleaning = _feat_eng = None

def _ensure_modules():
    global _cleaning, _feat_eng
    if _cleaning is None:
        _cleaning = _load_module("cleaning", os.path.join(SRC_DIR, "01_data_cleaning.py"))
    if _feat_eng is None:
        _feat_eng = _load_module("features", os.path.join(SRC_DIR, "02_feature_engineering.py"))

def _classify(p):
    if p >= HIGH_RISK_THRESHOLD: return "HIGH"
    if p >= MED_RISK_THRESHOLD:  return "MEDIUM"
    return "LOW"

def run_pipeline(csv_path: str) -> pd.DataFrame:
    _ensure_modules()
    df_raw = pd.read_csv(csv_path)
    rename_map = {
        "Transaction_Amount (in Million)": "Transaction_Amount",
        "Account_Balance (in Million)":    "Account_Balance",
        "Avg_Transaction_Amount (in Million)": "Avg_Transaction_Amount",
        "Max_Transaction_Last_24h (in Million)": "Max_Transaction_Last_24h",
        "Is_International_Transaction": "Is_International",
        "Unusual_Time_Transaction": "Unusual_Time",
    }
    df = df_raw.rename(columns=rename_map).copy()
    for col in df.select_dtypes(include=[np.number]).columns:
        df[col] = df[col].fillna(df[col].median())
    for col in df.select_dtypes(include=["object", "string"]).columns:
        if col != "Fraud_Label":
            mode = df[col].mode()
            df[col] = df[col].fillna(mode[0] if len(mode) > 0 else "")
    df["Transaction_Date"] = pd.to_datetime(df["Transaction_Date"], errors="coerce")
    df["Month"]     = df["Transaction_Date"].dt.month.fillna(1).astype(int)
    df["DayOfWeek"] = df["Transaction_Date"].dt.dayofweek.fillna(0).astype(int)
    def ph(t):
        try: return int(str(t).split(":")[0])
        except: return 12
    df["Hour"] = (df["Transaction_Time"].apply(ph)
                  if "Transaction_Time" in df.columns
                  else pd.Series([12] * len(df)))
    for col in ["Is_International", "Is_New_Merchant", "Unusual_Time"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().str.capitalize()
            df[col] = df[col].map({"Yes": "Yes", "No": "No"}).fillna("No")
    df = _feat_eng.engineer_features(df)
    model = joblib.load(MODEL_PATH)
    with open(META_PATH) as f:
        meta = json.load(f)
    feature_cols = meta["feature_cols"]
    for col in feature_cols:
        if col not in df.columns:
            df[col] = 0
    probs = model.predict_proba(df[feature_cols])[:, 1]

    # Dùng best_threshold nếu có trong meta
    threshold = meta.get("best_threshold", HIGH_RISK_THRESHOLD)

    result = df_raw.copy()
    result["Fraud_Probability"] = np.round(probs, 4)
    result["Risk_Level"]        = [_classify(p) for p in probs]
    return result


def _to_excel_bytes(df: pd.DataFrame) -> bytes:
    from openpyxl.styles import PatternFill
    from openpyxl import load_workbook
    risk_counts = df["Risk_Level"].value_counts()
    n = len(df)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All_Transactions", index=False)
        (df[df["Risk_Level"] == "HIGH"]
            .sort_values("Fraud_Probability", ascending=False)
            .to_excel(writer, sheet_name="HIGH_Risk_Alert", index=False))
        pd.DataFrame({
            "Risk_Level": ["LOW", "MEDIUM", "HIGH"],
            "Count":      [risk_counts.get(r, 0) for r in ["LOW","MEDIUM","HIGH"]],
            "Percentage": [f"{risk_counts.get(r,0)/n*100:.1f}%" for r in ["LOW","MEDIUM","HIGH"]],
        }).to_excel(writer, sheet_name="Summary", index=False)
    buf.seek(0)
    wb = load_workbook(buf)
    colors = {"LOW": "C8E6C9", "MEDIUM": "FFF9C4", "HIGH": "FFCDD2"}
    for sname in wb.sheetnames:
        ws = wb[sname]
        hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        if "Risk_Level" not in hdrs: continue
        rc = hdrs.index("Risk_Level") + 1
        for row in ws.iter_rows(min_row=2):
            v = row[rc - 1].value
            if v in colors:
                row[rc - 1].fill = PatternFill("solid", fgColor=colors[v])
    buf2 = io.BytesIO()
    wb.save(buf2)
    buf2.seek(0)
    return buf2.getvalue()


def _build_html_report(df_result: pd.DataFrame, filename: str) -> str:
    """Tạo báo cáo HTML trực tiếp trong scheduler (không phụ thuộc email_alert.py)."""
    df_high = df_result[df_result["Risk_Level"] == "HIGH"].sort_values("Fraud_Probability", ascending=False)
    n_high  = len(df_high)
    n_med   = (df_result["Risk_Level"] == "MEDIUM").sum()
    n_low   = (df_result["Risk_Level"] == "LOW").sum()
    n_total = len(df_result)
    now     = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    pct_high = n_high / n_total * 100 if n_total else 0
    pct_med  = n_med  / n_total * 100 if n_total else 0
    pct_low  = n_low  / n_total * 100 if n_total else 0

    show_cols = [c for c in [
        "Transaction_ID", "Transaction_Amount", "Transaction_Type",
        "Merchant_Category", "Card_Type", "Is_International",
        "Unusual_Time", "Fraud_Probability", "Risk_Level",
    ] if c in df_high.columns]

    headers_html = "".join(f"<th>{c.replace('_',' ')}</th>" for c in show_cols)
    rows_html = ""
    for _, row in df_high[show_cols].head(500).iterrows():
        prob = float(row.get("Fraud_Probability", 0))
        cells = ""
        for col in show_cols:
            val = row[col]
            if col == "Fraud_Probability":
                cells += f"<td><div class='pb'><div class='pf' style='width:{prob*100:.0f}%'></div></div><span class='pt'>{prob*100:.1f}%</span></td>"
            elif col == "Risk_Level":
                cells += "<td><span class='badge'>🔴 HIGH</span></td>"
            elif col == "Transaction_Amount":
                cells += f"<td class='mono'>${float(val):,.2f}M</td>"
            else:
                cells += f"<td>{val}</td>"
        rows_html += f"<tr>{cells}</tr>"

    return f"""<!DOCTYPE html>
<html lang="vi"><head><meta charset="utf-8">
<title>FraudShield Report — {filename}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:#f0f4f8;color:#1e293b}}
.header{{background:linear-gradient(135deg,#0f1117,#1e293b);padding:28px 40px}}
.logo{{font-size:26px;font-weight:800;color:#4fc3f7}}.sub{{color:#94a3b8;font-size:13px;margin-top:4px}}
.banner{{background:#fee2e2;border-left:5px solid #dc2626;padding:18px 40px}}
.bt{{color:#b91c1c;font-size:18px;font-weight:700}}.bs{{color:#7f1d1d;font-size:13px;margin-top:4px}}
.content{{padding:32px 40px;max-width:1400px;margin:0 auto}}
.kpi-row{{display:flex;gap:16px;margin-bottom:28px;flex-wrap:wrap}}
.kpi{{background:white;border-radius:12px;padding:20px 28px;flex:1;min-width:140px;box-shadow:0 1px 4px rgba(0,0,0,.08);text-align:center}}
.kv{{font-size:32px;font-weight:800}}.kl{{font-size:11px;color:#64748b;font-weight:600;text-transform:uppercase;letter-spacing:.08em;margin-top:6px}}
.red{{color:#dc2626}}.yellow{{color:#d97706}}.green{{color:#16a34a}}.dark{{color:#0f172a}}
.dist{{background:white;border-radius:12px;padding:20px 24px;box-shadow:0 1px 4px rgba(0,0,0,.08);margin-bottom:28px}}
.dist-bar{{display:flex;height:24px;border-radius:8px;overflow:hidden;gap:2px}}
.seg-h{{background:#dc2626;height:100%}}.seg-m{{background:#f59e0b;height:100%}}.seg-l{{background:#22c55e;height:100%}}
.legend{{display:flex;gap:20px;margin-top:10px;font-size:12px;color:#64748b}}
.dot{{width:10px;height:10px;border-radius:50%;display:inline-block;margin-right:5px;vertical-align:middle}}
.tcard{{background:white;border-radius:12px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.08);margin-bottom:24px}}
.ttitle{{padding:16px 20px;border-bottom:1px solid #f1f5f9;font-weight:700;font-size:14px}}
.twrap{{overflow-x:auto}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
th{{padding:10px 14px;background:#1e293b;color:white;text-align:left;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.05em;white-space:nowrap}}
td{{padding:9px 14px;border-bottom:1px solid #f1f5f9}}
tr:hover td{{background:#f8fafc}}
.mono{{font-family:monospace}}
.badge{{background:#fee2e2;color:#b91c1c;padding:3px 10px;border-radius:20px;font-weight:700;font-size:11px}}
.pb{{display:inline-block;width:70px;height:7px;background:#fee2e2;border-radius:4px;overflow:hidden;vertical-align:middle;margin-right:6px}}
.pf{{height:100%;background:#dc2626;border-radius:4px}}.pt{{color:#dc2626;font-weight:700;vertical-align:middle}}
.footer{{background:#e2e8f0;padding:20px 40px;text-align:center;color:#94a3b8;font-size:12px;margin-top:40px}}
.tag{{display:inline-block;background:#dbeafe;color:#1d4ed8;font-size:11px;font-weight:600;padding:4px 12px;border-radius:20px;margin-bottom:16px}}
</style></head><body>
<div class="header"><div class="logo">🛡️ FraudShield</div><div class="sub">Banking Fraud Detection · Automated Report</div></div>
<div class="banner"><div class="bt">⚠️ Phát hiện {n_high} giao dịch nghi ngờ gian lận</div><div class="bs">File: <strong>{filename}</strong> · {now}</div></div>
<div class="content">
<div class="kpi-row">
  <div class="kpi"><div class="kv red">{n_high}</div><div class="kl">🔴 HIGH Risk</div></div>
  <div class="kpi"><div class="kv yellow">{n_med}</div><div class="kl">🟡 MEDIUM Risk</div></div>
  <div class="kpi"><div class="kv green">{n_low}</div><div class="kl">🟢 LOW Risk</div></div>
  <div class="kpi"><div class="kv dark">{n_total:,}</div><div class="kl">Tổng GD</div></div>
</div>
<div class="dist">
  <div style="font-weight:700;font-size:14px;margin-bottom:14px">Phân bổ mức độ rủi ro</div>
  <div class="dist-bar">
    <div class="seg-h" style="width:{pct_high:.1f}%"></div>
    <div class="seg-m" style="width:{pct_med:.1f}%"></div>
    <div class="seg-l" style="width:{pct_low:.1f}%"></div>
  </div>
  <div class="legend">
    <span><span class="dot" style="background:#dc2626"></span>HIGH {pct_high:.1f}%</span>
    <span><span class="dot" style="background:#f59e0b"></span>MEDIUM {pct_med:.1f}%</span>
    <span><span class="dot" style="background:#22c55e"></span>LOW {pct_low:.1f}%</span>
  </div>
</div>
<div class="tcard">
  <div class="ttitle">🔴 Danh sách HIGH Risk ({n_high} giao dịch)</div>
  <div class="twrap"><table><thead><tr>{headers_html}</tr></thead><tbody>{rows_html}</tbody></table></div>
</div>
<div class="tag">📄 Báo cáo tự động · FraudShield · {now}</div>
</div>
<div class="footer">FraudShield Automated Report · FAA4023 Final Assignment 2026<br>Tạo tự động bởi scheduler. Không cần internet.</div>
</body></html>"""


def process_file(csv_path: str):
    filename = os.path.basename(csv_path)
    log.info(f"{'─'*55}")
    log.info(f"[→] Phát hiện file mới: {filename}")
    try:
        log.info("[1] Chạy pipeline: cleaning → features → predict…")
        result = run_pipeline(csv_path)
        n_total = len(result)
        n_high  = (result["Risk_Level"] == "HIGH").sum()
        n_med   = (result["Risk_Level"] == "MEDIUM").sum()
        n_low   = (result["Risk_Level"] == "LOW").sum()
        log.info(f"    ✓ {n_total:,} giao dịch | 🔴 HIGH={n_high} | 🟡 MED={n_med} | 🟢 LOW={n_low}")

        # Excel
        log.info("[2] Xuất báo cáo Excel…")
        excel_bytes = _to_excel_bytes(result)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_name = f"fraud_report_{Path(filename).stem}_{ts}.xlsx"
        with open(os.path.join(OUTPUT_DIR, excel_name), "wb") as f:
            f.write(excel_bytes)
        log.info(f"    ✓ Lưu → output/{excel_name}")

        # HTML report — dùng absolute path, tạo trực tiếp tại đây
        log.info("[3] Xuất báo cáo HTML…")
        html_name = f"fraud_report_{Path(filename).stem}_{ts}.html"
        html_path = os.path.join(REPORTS_DIR, html_name)
        html_content = _build_html_report(result, filename)
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        log.info(f"    [✓] Xuất HTML → output/reports/{html_name}  ({n_high} HIGH risk)")

        # Archive
        archive_dest = os.path.join(ARCHIVE_PATH, f"{ts}_{filename}")
        shutil.move(csv_path, archive_dest)
        log.info(f"[4] Di chuyển → inbox/processed/{ts}_{filename}")
        log.info(f"[✓] Hoàn thành xử lý: {filename}")

    except Exception as e:
        log.error(f"[✗] Lỗi khi xử lý {filename}: {e}")
        import traceback
        log.error(traceback.format_exc())


def main():
    print("\n" + "═"*55)
    print("  🛡️  FraudShield Automated Scheduler v2")
    print("═"*55)
    if not os.path.exists(MODEL_PATH):
        print("\n  ✗ Model chưa được train!")
        print("  → Chạy: streamlit run app.py → Chạy Pipeline Đầy đủ\n")
        sys.exit(1)

    print(f"\n  ✓ Model     : {MODEL_PATH}")
    print(f"  ✓ Inbox     : {INBOX_PATH}")
    print(f"  ✓ Reports   : {REPORTS_DIR}")
    print(f"  ✓ Interval  : mỗi {SCAN_INTERVAL_SECONDS}s")
    print(f"  ✓ Log       : {LOG_PATH}")
    print(f"\n  → Thả file CSV vào: inbox/")
    print(f"  → Báo cáo HTML tại: output/reports/")
    print(f"  → Nhấn Ctrl+C để dừng\n")
    print("═"*55 + "\n")

    log.info("=== FraudShield Scheduler v2 khởi động ===")
    log.info(f"Inbox: {INBOX_PATH} | Reports: {REPORTS_DIR} | Interval: {SCAN_INTERVAL_SECONDS}s")

    scan_count = 0
    try:
        while True:
            scan_count += 1
            csv_files = sorted(Path(INBOX_PATH).glob("*.csv"), key=lambda p: p.stat().st_mtime)
            if csv_files:
                log.info(f"[Scan #{scan_count}] Tìm thấy {len(csv_files)} file CSV")
                for csv_path in csv_files:
                    process_file(str(csv_path))
            else:
                if scan_count <= 5 or scan_count % 10 == 0:
                    log.info(f"[Scan #{scan_count}] Không có file mới. Chờ {SCAN_INTERVAL_SECONDS}s…")
            time.sleep(SCAN_INTERVAL_SECONDS)
    except KeyboardInterrupt:
        print("\n")
        log.info("=== Scheduler dừng (Ctrl+C) ===")
        print("\n  Scheduler đã dừng.\n")

if __name__ == "__main__":
    main()