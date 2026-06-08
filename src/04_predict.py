"""
FraudShield — Bước 4: Dự đoán tự động trên file CSV mới
========================================================
Input : file CSV giao dịch mới (chưa có Fraud_Label)
Output: output/fraud_report.xlsx  (3 sheets)

Sử dụng: python 04_predict.py <input.csv>  [output.xlsx]
"""

import pandas as pd
import numpy as np
import os, sys, joblib, json

BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MODEL_PATH = os.path.join(BASE_DIR, "model", "fraud_model.pkl")
META_PATH  = os.path.join(BASE_DIR, "model", "model_meta.json")
REPORT_PATH = os.path.join(BASE_DIR, "output", "fraud_report.xlsx")

# Import các hàm từ script khác
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from src_01_cleaning import clean_data_df
from src_02_features import engineer_features


def log(msg): print(f"  {msg}")


def classify_risk(prob: float) -> str:
    if prob < 0.30:
        return "LOW"
    elif prob < 0.60:
        return "MEDIUM"
    else:
        return "HIGH"


def predict(input_csv: str,
            output_xlsx: str = REPORT_PATH,
            model_path: str = MODEL_PATH) -> pd.DataFrame:
    """
    Hàm dự đoán chính. Có thể gọi từ Streamlit.
    Trả về DataFrame kết quả đầy đủ.
    """
    print("\n" + "═" * 60)
    print("  FraudShield │ 04_predict.py")
    print("═" * 60)

    # ── 1. Load model + meta ──────────────────────────────────
    model = joblib.load(model_path)
    with open(META_PATH) as f:
        meta = json.load(f)
    feature_cols = meta["feature_cols"]
    log(f"[✓] Load model từ {os.path.basename(model_path)}")
    log(f"    AUC={meta['metrics']['AUC']}  Recall={meta['metrics']['Recall']}")

    # ── 2. Đọc + làm sạch data mới ───────────────────────────
    df_raw = pd.read_csv(input_csv)
    log(f"[✓] Đọc {len(df_raw):,} giao dịch từ {os.path.basename(input_csv)}")
    df_clean = clean_data_df(df_raw)
    df_feat  = engineer_features(df_clean)
    log("[✓] Làm sạch + tạo features tự động")

    # ── 3. Chuẩn bị feature matrix ───────────────────────────
    # Đảm bảo có đủ cột (thêm cột 0 nếu thiếu)
    for col in feature_cols:
        if col not in df_feat.columns:
            df_feat[col] = 0
    X = df_feat[feature_cols]

    # ── 4. Dự đoán ───────────────────────────────────────────
    probs  = model.predict_proba(X)[:, 1]
    df_raw = df_raw.head(len(probs)).copy()   # align length
    df_raw["Fraud_Probability"] = np.round(probs, 4)
    df_raw["Risk_Level"]        = [classify_risk(p) for p in probs]
    log("[✓] Dự đoán xong → Risk Score + Risk Level")

    # ── 5. Xuất Excel 3 sheets ───────────────────────────────
    _export_excel(df_raw, output_xlsx)
    log(f"[✓] Xuất báo cáo Excel → {os.path.basename(output_xlsx)}")
    print("═" * 60 + "\n")

    return df_raw


def _export_excel(df: pd.DataFrame, path: str):
    """Tạo Excel 3 sheets với format màu Risk Level."""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    risk_counts = df["Risk_Level"].value_counts()

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # Sheet 1: All transactions
        df.to_excel(writer, sheet_name="All_Transactions", index=False)

        # Sheet 2: HIGH risk only
        df_high = df[df["Risk_Level"] == "HIGH"].copy()
        df_high = df_high.sort_values("Fraud_Probability", ascending=False)
        df_high.to_excel(writer, sheet_name="HIGH_Risk_Alert", index=False)

        # Sheet 3: Summary
        summary = pd.DataFrame({
            "Risk_Level": ["LOW", "MEDIUM", "HIGH"],
            "Count":      [risk_counts.get("LOW", 0),
                           risk_counts.get("MEDIUM", 0),
                           risk_counts.get("HIGH", 0)],
            "Percentage": [
                f"{risk_counts.get('LOW',0)/len(df)*100:.1f}%",
                f"{risk_counts.get('MEDIUM',0)/len(df)*100:.1f}%",
                f"{risk_counts.get('HIGH',0)/len(df)*100:.1f}%",
            ],
        })
        summary.to_excel(writer, sheet_name="Summary", index=False)

    # ── Tô màu Risk Level bằng openpyxl ──────────────────────
    wb = load_workbook(path)
    colors = {"LOW": "C8E6C9", "MEDIUM": "FFF9C4", "HIGH": "FFCDD2"}

    for sheet_name in ["All_Transactions", "HIGH_Risk_Alert"]:
        ws = wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        risk_col = headers.index("Risk_Level") + 1 if "Risk_Level" in headers else None
        if not risk_col: continue
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            risk = row[risk_col - 1].value
            if risk in colors:
                fill = PatternFill("solid", fgColor=colors[risk])
                row[risk_col - 1].fill = fill

        # Auto-width columns
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)

    wb.save(path)


# ─────────────────────────────────────────────────────────────
# Helper: wrap clean + feature cho import (không ghi file)
# ─────────────────────────────────────────────────────────────

def _import_helper():
    """Tạo module tạm thời để 04_predict có thể import từ src."""
    pass


# Tạo aliases để 04_predict tự-import được khi chạy standalone
import importlib.util

def _load_module(name, path):
    spec = importlib.util.spec_from_file_location(name, path)
    mod  = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# Lazy load — chỉ chạy khi gọi từ command line
if __name__ == "__main__":
    _src_dir = os.path.dirname(os.path.abspath(__file__))

    _cleaning  = _load_module("cleaning", os.path.join(_src_dir, "01_data_cleaning.py"))
    _feat_eng  = _load_module("feat_eng", os.path.join(_src_dir, "02_feature_engineering.py"))

    # Gắn các hàm cần thiết vào namespace
    def clean_data_df(df):
        df = df.copy()
        rename_map = {
            "Transaction_Amount (in Million)": "Transaction_Amount",
            "Account_Balance (in Million)":    "Account_Balance",
            "Avg_Transaction_Amount (in Million)": "Avg_Transaction_Amount",
            "Max_Transaction_Last_24h (in Million)": "Max_Transaction_Last_24h",
            "Is_International_Transaction": "Is_International",
            "Unusual_Time_Transaction": "Unusual_Time",
        }
        df.rename(columns=rename_map, inplace=True)
        # Impute
        for col in df.select_dtypes(include=np.number).columns:
            df[col].fillna(df[col].median(), inplace=True)
        for col in df.select_dtypes(include="object").columns:
            if col != "Fraud_Label":
                mode = df[col].mode()
                df[col].fillna(mode[0] if len(mode) > 0 else "", inplace=True)
        # Parse time
        df["Transaction_Date"] = pd.to_datetime(df["Transaction_Date"], errors="coerce")
        df["Month"]     = df["Transaction_Date"].dt.month.fillna(1).astype(int)
        df["DayOfWeek"] = df["Transaction_Date"].dt.dayofweek.fillna(0).astype(int)
        def ph(t):
            try: return int(str(t).split(":")[0])
            except: return 12
        df["Hour"] = df["Transaction_Time"].apply(ph)
        for col in ["Is_International", "Is_New_Merchant", "Unusual_Time"]:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip().str.capitalize()
                df[col] = df[col].map({"Yes":"Yes","No":"No"}).fillna("No")
        return df

    from src_01_cleaning import clean_data_df  # noqa (replaced by local def above)
    from src_02_features  import engineer_features  # noqa

    inp  = sys.argv[1] if len(sys.argv) > 1 else os.path.join(BASE_DIR, "data", "FraudShield_Banking_Data.csv")
    out  = sys.argv[2] if len(sys.argv) > 2 else REPORT_PATH
    predict(inp, out)
else:
    # Khi import từ app.py — định nghĩa helper functions inline
    import importlib.util as _ilu

    def _get_cleaner():
        p = os.path.join(os.path.dirname(__file__), "01_data_cleaning.py")
        m = _ilu.spec_from_file_location("_c", p)
        mod = _ilu.module_from_spec(m); m.loader.exec_module(mod)
        return mod

    def _get_feater():
        p = os.path.join(os.path.dirname(__file__), "02_feature_engineering.py")
        m = _ilu.spec_from_file_location("_f", p)
        mod = _ilu.module_from_spec(m); m.loader.exec_module(mod)
        return mod
