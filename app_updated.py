"""
FraudShield — Dashboard Streamlit v5
5-tab flow: Upload & Clean → Predict → Overview → Analysis → Search
"""

import streamlit as st
import pandas as pd
import numpy as np
import os, json, joblib, io, sys, math
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from datetime import datetime
from pathlib import Path

st.set_page_config(
    page_title="FraudShield Analytics",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded",
)

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
DATA_RAW    = os.path.join(BASE_DIR, "data", "FraudShield_Banking_Data.csv")
DATA_CLEAN  = os.path.join(BASE_DIR, "data", "cleaned_data.csv")
DATA_FEAT   = os.path.join(BASE_DIR, "data", "featured_data.csv")
MODEL_PATH  = os.path.join(BASE_DIR, "model", "fraud_model.pkl")
META_PATH   = os.path.join(BASE_DIR, "model", "model_meta.json")
OUTPUT_DIR  = os.path.join(BASE_DIR, "output")
REPORTS_DIR = os.path.join(BASE_DIR, "output", "reports")
SRC_DIR     = os.path.join(BASE_DIR, "src")
INBOX_DIR   = os.path.join(BASE_DIR, "inbox")
ARCHIVE_DIR = os.path.join(BASE_DIR, "inbox", "processed")
sys.path.insert(0, SRC_DIR)
sys.path.insert(0, BASE_DIR)

# ═══════════════════════════════════════════════════════════
# DESIGN SYSTEM — Revolut Indigo Dark
# ═══════════════════════════════════════════════════════════
C_BG       = "#090714"
C_BG2      = "#121022"
C_BG3      = "#1B1733"
C_BORDER   = "#34284F"
C_ACCENT   = "#D946EF"
C_ACCENT2  = "#7C3AED"
C_ACCENT_L = "#F0ABFC"
C_TEXT     = "#F1F5F9"
C_MUTED    = "#94A3B8"
C_HIGH     = "#FF4D9D"
C_MED      = "#C084FC"
C_LOW      = "#8B5CF6"
C_CHART1   = "#D946EF"
C_CHART2   = "#A855F7"
C_CHART3   = "#6D5D9A"

# Required columns for the model
REQUIRED_COLS = [
    "Transaction_Amount", "Account_Balance", "Transaction_Type",
    "Merchant_Category", "Card_Type", "Is_International",
    "Is_New_Merchant", "Unusual_Time", "Distance_From_Home",
    "Daily_Transaction_Count", "Weekly_Transaction_Count",
    "Avg_Transaction_Amount", "Max_Transaction_Last_24h",
    "Failed_Transaction_Count", "Previous_Fraud_Count",
    "Transaction_Date", "Transaction_Time",
]

RENAME_MAP = {
    "Transaction_Amount (in Million)":      "Transaction_Amount",
    "Account_Balance (in Million)":         "Account_Balance",
    "Avg_Transaction_Amount (in Million)":  "Avg_Transaction_Amount",
    "Max_Transaction_Last_24h (in Million)":"Max_Transaction_Last_24h",
    "Is_International_Transaction":         "Is_International",
    "Unusual_Time_Transaction":             "Unusual_Time",
}

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500;600&display=swap');

html, body, [class*="css"] {{
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    background: {C_BG} !important;
    color: {C_TEXT} !important;
}}
.stApp {{ background: {C_BG} !important; }}

::-webkit-scrollbar {{ width: 5px; height: 5px; }}
::-webkit-scrollbar-track {{ background: {C_BG2}; }}
::-webkit-scrollbar-thumb {{ background: {C_BORDER}; border-radius: 4px; }}
::-webkit-scrollbar-thumb:hover {{ background: {C_ACCENT}; }}

[data-testid="stSidebar"] {{
    background: {C_BG2} !important;
    border-right: 1px solid {C_BORDER} !important;
}}
[data-testid="stSidebar"] * {{
    color: {C_TEXT} !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
}}

/* ── TABS ── */
.stTabs [data-baseweb="tab-list"] {{
    background: transparent;
    border-bottom: 1px solid {C_BORDER};
    gap: 0; padding: 0;
}}
.stTabs [data-baseweb="tab"] {{
    background: transparent !important;
    color: {C_MUTED} !important;
    border-radius: 0 !important;
    font-weight: 600 !important;
    font-size: 12px !important;
    padding: 14px 20px !important;
    border: none !important;
    border-bottom: 2px solid transparent !important;
    transition: all 0.2s ease;
    letter-spacing: 0.04em;
    text-transform: uppercase;
}}
.stTabs [data-baseweb="tab"]:hover {{
    color: {C_TEXT} !important;
    background: {C_BG3} !important;
}}
.stTabs [aria-selected="true"] {{
    color: {C_ACCENT_L} !important;
    border-bottom: 2px solid {C_ACCENT} !important;
    background: transparent !important;
    font-weight: 800 !important;
}}

/* ── KPI CARDS ── */
.kpi-card {{
    background: {C_BG2};
    border: 1px solid {C_BORDER};
    border-radius: 16px;
    padding: 24px 22px 20px;
    position: relative;
    overflow: hidden;
    transition: border-color 0.2s, transform 0.2s;
}}
.kpi-card:hover {{ border-color: {C_ACCENT}60; transform: translateY(-1px); }}
.kpi-card::after {{
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0; height: 1px;
    background: linear-gradient(90deg, transparent, {C_ACCENT}80, transparent);
}}
.kpi-value {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 2.1rem; font-weight: 600;
    color: {C_TEXT}; line-height: 1;
    margin-bottom: 10px; letter-spacing: -1.5px;
}}
.kpi-label {{
    font-size: 0.67rem; color: {C_MUTED};
    font-weight: 700; text-transform: uppercase; letter-spacing: 0.14em;
}}
.kpi-red    .kpi-value {{ color: {C_HIGH}; }}
.kpi-amber  .kpi-value {{ color: {C_MED};  }}
.kpi-green  .kpi-value {{ color: {C_LOW};  }}
.kpi-indigo .kpi-value {{ color: {C_ACCENT_L}; }}

/* ── SECTION HEADER ── */
.sec-hdr {{
    font-size: 0.7rem; font-weight: 800; color: {C_MUTED};
    text-transform: uppercase; letter-spacing: 0.16em;
    padding-bottom: 10px; margin: 28px 0 16px 0;
    border-bottom: 1px solid {C_BORDER};
    display: flex; align-items: center; gap: 8px;
}}
.sec-hdr::before {{
    content: ''; display: inline-block;
    width: 3px; height: 13px;
    background: {C_ACCENT}; border-radius: 2px; flex-shrink: 0;
}}

/* ── PAGE TITLE ── */
.page-title {{
    font-size: 1.45rem; font-weight: 800; color: {C_TEXT};
    letter-spacing: -0.4px; margin-bottom: 4px; line-height: 1.2;
}}
.page-sub {{
    font-size: 13px; color: {C_MUTED};
    font-weight: 400; margin-bottom: 28px;
}}

/* ── STEP INDICATOR ── */
.step-flow {{
    display: flex; align-items: center; gap: 0;
    margin-bottom: 28px; overflow-x: auto;
}}
.step-item {{
    display: flex; align-items: center; gap: 10px;
    background: {C_BG2}; border: 1px solid {C_BORDER};
    border-radius: 10px; padding: 10px 16px;
    font-size: 12px; font-weight: 600; color: {C_MUTED};
    white-space: nowrap; flex-shrink: 0;
}}
.step-item.active {{
    border-color: {C_ACCENT}; color: {C_ACCENT_L};
    background: {C_ACCENT}12;
}}
.step-item.done {{
    border-color: {C_LOW}40; color: {C_LOW};
    background: {C_LOW}08;
}}
.step-arrow {{ color: {C_BORDER}; font-size: 16px; padding: 0 6px; flex-shrink: 0; }}

/* ── COLUMN MAPPING TABLE ── */
.col-row {{
    display: flex; align-items: center; gap: 12px;
    padding: 10px 14px; border-radius: 8px;
    margin-bottom: 4px; font-size: 13px;
    border: 1px solid {C_BORDER}; background: {C_BG2};
}}
.col-match   {{ border-color: {C_LOW}30;   background: {C_LOW}06;   }}
.col-missing {{ border-color: {C_HIGH}30;  background: {C_HIGH}06;  }}
.col-rename  {{ border-color: {C_MED}30;   background: {C_MED}06;   }}

/* ── PIPELINE STEPS ── */
.pipe-step {{
    display: flex; align-items: center; gap: 10px;
    background: {C_BG}; border-radius: 10px;
    padding: 10px 14px; margin: 5px 0;
    border: 1px solid {C_BORDER};
    font-size: 0.78rem; color: {C_MUTED};
    font-family: 'JetBrains Mono', monospace;
}}
.pipe-step.done {{ border-color: {C_ACCENT}50; background: {C_ACCENT}08; color: {C_TEXT}; }}
.dot {{ width: 8px; height: 8px; border-radius: 50%; flex-shrink: 0; }}
.dot-done {{ background: {C_ACCENT}; box-shadow: 0 0 8px {C_ACCENT}60; }}
.dot-idle {{ background: {C_BORDER}; }}

/* ── BUTTONS ── */
.stButton > button {{
    background: linear-gradient(135deg, {C_ACCENT}, {C_ACCENT2}) !important;
    color: white !important; font-weight: 700 !important;
    font-size: 13px !important; border: none !important;
    border-radius: 10px !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    padding: 10px 20px !important;
    transition: all 0.15s ease !important;
    box-shadow: 0 4px 14px {C_ACCENT}40 !important;
}}
.stButton > button:hover {{
    background: {C_ACCENT2} !important;
    transform: translateY(-1px) !important;
}}

/* ── DATAFRAME ── */
[data-testid="stDataFrame"] {{
    border-radius: 12px !important;
    overflow: hidden !important;
    border: 1px solid {C_BORDER} !important;
}}

/* ── INPUTS ── */
[data-baseweb="select"] > div {{
    background: {C_BG2} !important; border-color: {C_BORDER} !important;
    border-radius: 10px !important; color: {C_TEXT} !important;
}}
[data-testid="stFileUploader"] {{
    background: {C_BG2}; border: 1px dashed {C_BORDER};
    border-radius: 12px; padding: 8px;
}}
[data-testid="stMetric"] {{
    background: {C_BG2}; border: 1px solid {C_BORDER};
    border-radius: 12px; padding: 14px 16px !important;
}}
[data-testid="stMetricValue"] {{
    font-family: 'JetBrains Mono', monospace;
    color: {C_TEXT} !important; font-size: 1.4rem !important;
}}
[data-testid="stMetricLabel"] {{
    color: {C_MUTED} !important; font-size: 10px !important;
    font-weight: 700 !important; text-transform: uppercase; letter-spacing: 0.1em;
}}
.stAlert {{ border-radius: 10px !important; }}
hr {{ border-color: {C_BORDER} !important; opacity: 0.5; }}
</style>
""", unsafe_allow_html=True)

# Final UI polish: giảm cảm giác "neon/dashboard game", ưu tiên corporate analytics.
st.markdown(f"""
<style>
.kpi-card {{
    border-radius: 14px !important;
    box-shadow: none !important;
    padding: 20px 20px 18px !important;
}}
.kpi-card:hover {{ transform: none !important; }}
.kpi-card::after {{ display:none !important; }}
.kpi-value {{ font-size: 1.85rem !important; letter-spacing: -1px !important; }}
.page-title {{ font-size: 1.35rem !important; }}
.page-sub {{ margin-bottom: 20px !important; }}
.sec-hdr {{ margin-top: 24px !important; }}
.step-item {{ border-radius: 8px !important; padding: 9px 13px !important; }}
.stRadio [role="radiogroup"] {{ gap: 10px; }}
[data-testid="stDataFrame"] {{ box-shadow: none !important; }}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────
def chart_fig(w=10, h=5):
    fig, ax = plt.subplots(figsize=(w, h))
    fig.patch.set_facecolor(C_BG2); ax.set_facecolor(C_BG2)
    ax.tick_params(colors=C_MUTED, labelsize=9, length=0)
    for sp in ax.spines.values(): sp.set_visible(False)
    ax.xaxis.label.set_color(C_MUTED); ax.yaxis.label.set_color(C_MUTED)
    ax.title.set_color(C_TEXT)
    ax.grid(axis='y', color=C_BORDER, linewidth=0.5, alpha=0.7)
    ax.set_axisbelow(True)
    return fig, ax

dark_fig = chart_fig

def add_bar_labels(ax, bars, fmt="{:.0f}", color=None):
    for bar in bars:
        h = bar.get_height()
        if h > 0:
            ax.text(bar.get_x() + bar.get_width()/2, h + h*0.015,
                    fmt.format(h), ha='center', va='bottom',
                    fontsize=8, color=color or C_MUTED,
                    fontweight='600', fontfamily='monospace')

def kpi(value, label, variant=""):
    return (f'<div class="kpi-card kpi-{variant}">'
            f'<div class="kpi-value">{value}</div>'
            f'<div class="kpi-label">{label}</div></div>')

def page_header(title, subtitle=""):
    st.markdown(
        f'<div class="page-title">{title}</div>'
        f'<div class="page-sub">{subtitle}</div>',
        unsafe_allow_html=True)

def sec_header(title):
    st.markdown(f'<div class="sec-hdr">{title}</div>', unsafe_allow_html=True)

def step_flow(steps, active_idx):
    """Hiển thị thanh tiến trình các bước."""
    html = '<div class="step-flow">'
    for i, (icon, label) in enumerate(steps):
        cls = "active" if i == active_idx else ("done" if i < active_idx else "")
        check = "✓ " if i < active_idx else ""
        html += f'<div class="step-item {cls}">{icon} {check}{label}</div>'
        if i < len(steps) - 1:
            html += '<span class="step-arrow">›</span>'
    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────
# PIPELINE & PREDICT FUNCTIONS
# ─────────────────────────────────────────────────────────
def run_full_pipeline():
    import importlib.util
    def lm(name, path):
        s = importlib.util.spec_from_file_location(name, path)
        m = importlib.util.module_from_spec(s); s.loader.exec_module(m); return m
    status = {}
    try:
        lm("c", os.path.join(SRC_DIR, "01_data_cleaning.py")).clean_data(DATA_RAW, DATA_CLEAN)
        status["clean"] = True
    except Exception as e: status["clean"] = str(e); return status
    try:
        lm("f", os.path.join(SRC_DIR, "02_feature_engineering.py")).run(DATA_CLEAN, DATA_FEAT)
        status["features"] = True
    except Exception as e: status["features"] = str(e); return status
    try:
        lm("t", os.path.join(SRC_DIR, "03_train_model.py")).train(DATA_FEAT)
        status["train"] = True
    except Exception as e: status["train"] = str(e)
    return status

def classify_risk(p):
    return "HIGH" if p >= 0.6 else ("MEDIUM" if p >= 0.3 else "LOW")

def clean_uploaded_df(df_raw: pd.DataFrame, col_map: dict) -> pd.DataFrame:
    """Làm sạch DataFrame đã được map cột."""
    df = df_raw.rename(columns=col_map).copy()
    # Impute numeric
    for col in df.select_dtypes(include=np.number).columns:
        df[col] = df[col].fillna(df[col].median())
    # Impute categorical
    for col in df.select_dtypes(include=["object", "string"]).columns:
        if col != "Fraud_Label":
            mode = df[col].mode()
            df[col] = df[col].fillna(mode[0] if len(mode) > 0 else "Unknown")
    # Parse date/time
    if "Transaction_Date" in df.columns:
        df["Transaction_Date"] = pd.to_datetime(df["Transaction_Date"], errors="coerce")
        df["Month"]     = df["Transaction_Date"].dt.month.fillna(1).astype(int)
        df["DayOfWeek"] = df["Transaction_Date"].dt.dayofweek.fillna(0).astype(int)
    if "Transaction_Time" in df.columns:
        ph = lambda t: int(str(t).split(":")[0]) if ":" in str(t) else 12
        df["Hour"] = df["Transaction_Time"].apply(ph)
    # Normalize Yes/No
    for col in ["Is_International", "Is_New_Merchant", "Unusual_Time"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().str.capitalize()
            df[col] = df[col].map({"Yes": "Yes", "No": "No"}).fillna("No")
    return df

def predict_from_clean(df_clean: pd.DataFrame) -> pd.DataFrame:
    """Chạy feature engineering + predict trên data đã clean."""
    import importlib.util
    def lm(name, path):
        s = importlib.util.spec_from_file_location(name, path)
        m = importlib.util.module_from_spec(s); s.loader.exec_module(m); return m
    m2 = lm("f", os.path.join(SRC_DIR, "02_feature_engineering.py"))
    df = m2.engineer_features(df_clean.copy())
    model = joblib.load(MODEL_PATH)
    with open(META_PATH) as f: meta = json.load(f)
    fc = meta["feature_cols"]
    for col in fc:
        if col not in df.columns: df[col] = 0
    probs = model.predict_proba(df[fc])[:, 1]
    result = df_clean.copy()
    result["Fraud_Probability"] = np.round(probs, 4)
    result["Risk_Level"] = [classify_risk(p) for p in probs]
    return result

def df_to_excel_bytes(df):
    from openpyxl.styles import PatternFill; from openpyxl import load_workbook
    rc_cnt = df["Risk_Level"].value_counts() if "Risk_Level" in df.columns else {}
    n = len(df); buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="All_Transactions", index=False)
        if "Risk_Level" in df.columns:
            df[df["Risk_Level"]=="HIGH"].sort_values("Fraud_Probability", ascending=False)\
              .to_excel(w, sheet_name="HIGH_Risk_Alert", index=False)
            pd.DataFrame({
                "Risk_Level": ["LOW","MEDIUM","HIGH"],
                "Count": [rc_cnt.get(r,0) for r in ["LOW","MEDIUM","HIGH"]],
                "Percentage": [f"{rc_cnt.get(r,0)/n*100:.1f}%" for r in ["LOW","MEDIUM","HIGH"]],
            }).to_excel(w, sheet_name="Summary", index=False)
    buf.seek(0); wb = load_workbook(buf)
    colors = {"LOW":"C8E6C9","MEDIUM":"FFF9C4","HIGH":"FFCDD2"}
    for sn in wb.sheetnames:
        ws = wb[sn]; hdrs = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
        if "Risk_Level" not in hdrs: continue
        rc = hdrs.index("Risk_Level")+1
        for row in ws.iter_rows(min_row=2):
            v = row[rc-1].value
            if v in colors: row[rc-1].fill = PatternFill("solid", fgColor=colors[v])
    buf2 = io.BytesIO(); wb.save(buf2); buf2.seek(0); return buf2.getvalue()

CHART_COLORS = [C_CHART1, C_CHART2, "#A78BFA", "#C4B5FD", C_CHART3]


def style_risk_cell(v):
    if v == "HIGH": return f"background:{C_HIGH}20;color:{C_HIGH};font-weight:800"
    if v == "MEDIUM": return f"background:{C_MED}20;color:{C_MED};font-weight:800"
    if v == "LOW": return f"background:{C_LOW}20;color:{C_LOW};font-weight:800"
    return ""

def risk_priority_label(row):
    tid = str(row.get("Transaction_ID", row.name))
    prob = float(row.get("Fraud_Probability", 0)) * 100
    amount = row.get("Transaction_Amount", "?")
    tx = row.get("Transaction_Type", "?")
    mc = row.get("Merchant_Category", "?")
    intl = row.get("Is_International", "?")
    unusual = row.get("Unusual_Time", "?")
    return f"ID {tid} · {prob:.1f}% · {tx} · {mc} · Amt {amount} · Intl {intl} · Unusual {unusual}"

def insight_card(title, value, subtitle, color=C_ACCENT_L):
    """Insight card đồng bộ hồng-tím, cố định chiều cao để các card bằng nhau."""
    html = f"""
    <div style="
        height:132px;
        background:linear-gradient(135deg, rgba(217,70,239,0.13), rgba(124,58,237,0.08));
        border:1px solid {C_BORDER};
        border-top:1px solid {color}70;
        border-left:3px solid {color};
        border-radius:16px;
        padding:18px 20px;
        box-sizing:border-box;
        display:flex;
        flex-direction:column;
        justify-content:space-between;
        box-shadow:0 10px 28px rgba(124,58,237,0.10);
    ">
        <div style="font-size:10px;color:{C_MUTED};font-weight:800;text-transform:uppercase;
                    letter-spacing:0.14em;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">{title}</div>
        <div style="font-family:'JetBrains Mono',monospace;font-size:1.45rem;font-weight:800;
                    color:{C_TEXT};line-height:1;letter-spacing:-0.04em;">{value}</div>
        <div style="font-size:12px;color:{C_MUTED};line-height:1.45;min-height:34px;">{subtitle}</div>
    </div>"""
    st.markdown(html, unsafe_allow_html=True)

def group_risk_summary(df, col):
    g = df.groupby(col).agg(
        Total=("Risk_Level", "count"),
        High=("Risk_Level", lambda x: (x == "HIGH").sum()),
        Medium=("Risk_Level", lambda x: (x == "MEDIUM").sum()),
        Avg_Prob=("Fraud_Probability", "mean"),
    ).reset_index()
    g["High_Rate"] = np.where(g["Total"] > 0, g["High"] / g["Total"] * 100, 0)
    g["Avg_Prob_%"] = g["Avg_Prob"] * 100
    return g.sort_values(["High_Rate", "High"], ascending=False)

def risk_counts_from_level(df):
    """Một nguồn duy nhất cho toàn bộ KPI/chart để tránh lệch số giữa các tab."""
    rc = df["Risk_Level"].value_counts() if "Risk_Level" in df.columns else pd.Series(dtype=int)
    return {
        "HIGH": int(rc.get("HIGH", 0)),
        "MEDIUM": int(rc.get("MEDIUM", 0)),
        "LOW": int(rc.get("LOW", 0)),
    }


# ─────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f"""
    <div style="padding:24px 16px 20px;border-bottom:1px solid {C_BORDER};">
        <div style="display:flex;align-items:center;gap:12px;margin-bottom:12px;">
            <div style="width:36px;height:36px;border-radius:10px;
                        background:linear-gradient(135deg,{C_ACCENT},{C_ACCENT2});
                        display:flex;align-items:center;justify-content:center;
                        font-size:18px;box-shadow:0 4px 12px {C_ACCENT}40;">🛡️</div>
            <div>
                <div style="font-size:15px;font-weight:800;color:{C_TEXT};letter-spacing:-0.3px;">
                    FraudShield</div>
                <div style="font-size:9px;color:{C_MUTED};font-weight:600;
                            letter-spacing:0.12em;text-transform:uppercase;margin-top:2px;">
                    Analytics Platform</div>
            </div>
        </div>
        <div style="display:flex;gap:6px;align-items:center;">
            <div style="width:6px;height:6px;border-radius:50%;background:{C_LOW};
                        box-shadow:0 0 6px {C_LOW}80;"></div>
            <span style="font-size:10px;color:{C_MUTED};">FAA4023 · Banking Fraud Detection</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Session status
    has_data    = "clean_df" in st.session_state
    has_predict = "predict_result" in st.session_state
    st.markdown(f"""
    <div style="padding:14px 16px 10px;">
        <div style="font-size:9px;font-weight:800;color:{C_MUTED};
                    letter-spacing:0.14em;text-transform:uppercase;margin-bottom:10px;">
            TRẠNG THÁI SESSION</div>
        <div style="display:flex;flex-direction:column;gap:6px;">
            <div style="display:flex;align-items:center;gap:8px;font-size:12px;">
                <div style="width:8px;height:8px;border-radius:50%;
                            background:{'#10B981' if has_data else C_BORDER};
                            box-shadow:{'0 0 6px #10B98180' if has_data else 'none'};
                            flex-shrink:0;"></div>
                <span style="color:{'#10B981' if has_data else C_MUTED};">
                    {'✓ Dữ liệu đã tải' if has_data else 'Chưa upload dữ liệu'}</span>
            </div>
            <div style="display:flex;align-items:center;gap:8px;font-size:12px;">
                <div style="width:8px;height:8px;border-radius:50%;
                            background:{'#10B981' if has_predict else C_BORDER};
                            box-shadow:{'0 0 6px #10B98180' if has_predict else 'none'};
                            flex-shrink:0;"></div>
                <span style="color:{'#10B981' if has_predict else C_MUTED};">
                    {'✓ Đã dự đoán' if has_predict else 'Chưa dự đoán'}</span>
            </div>
        </div>
    </div>
    <div style="padding:0 16px 10px;border-bottom:1px solid {C_BORDER};"></div>
    """, unsafe_allow_html=True)

    # Pipeline
    st.markdown(f"""
    <div style="padding:14px 16px 8px;font-size:9px;font-weight:800;
                color:{C_MUTED};letter-spacing:0.14em;text-transform:uppercase;">
        ML PIPELINE</div>""", unsafe_allow_html=True)
    model_ready = os.path.exists(MODEL_PATH)
    for path, label in [
        (DATA_CLEAN, "01_data_cleaning.py"),
        (DATA_FEAT,  "02_feature_engineering.py"),
        (MODEL_PATH, "03_train_model.py"),
    ]:
        done = os.path.exists(path)
        st.markdown(
            f'<div class="pipe-step {"done" if done else ""}">'
            f'<div class="dot {"dot-done" if done else "dot-idle"}"></div>{label}</div>',
            unsafe_allow_html=True)
    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
    if st.button("▶  RUN FULL PIPELINE", use_container_width=True, type="primary"):
        with st.spinner("Running pipeline…"):
            status = run_full_pipeline()
        if all(v is True for v in status.values()):
            st.success("Pipeline completed!"); st.rerun()
        else:
            for k, v in status.items():
                if v is not True: st.error(f"Error in {k}: {v}")

    # Model metrics
    if model_ready:
        with open(META_PATH) as f: meta = json.load(f)
        m = meta["metrics"]
        st.markdown(f"""
        <div style="padding:14px 16px 8px;font-size:9px;font-weight:800;
                    color:{C_MUTED};letter-spacing:0.14em;text-transform:uppercase;
                    margin-top:8px;">MODEL PERFORMANCE</div>""", unsafe_allow_html=True)
        def mbar(label, val, g, b):
            c = C_LOW if val>=g else (C_MED if val>=b else C_HIGH)
            return f"""<div style="margin-bottom:10px;padding:0 16px;">
                <div style="display:flex;justify-content:space-between;margin-bottom:4px;">
                    <span style="font-size:10px;color:{C_MUTED};font-weight:600;">{label}</span>
                    <span style="font-family:'JetBrains Mono',monospace;font-size:11px;
                                 font-weight:700;color:{c};">{val:.3f}</span>
                </div>
                <div style="background:{C_BORDER};border-radius:3px;height:3px;overflow:hidden;">
                    <div style="width:{val*100:.0f}%;height:100%;background:{c};border-radius:3px;"></div>
                </div></div>"""
        st.markdown(
            mbar("AUC-ROC",   m["AUC"],       0.80, 0.65) +
            mbar("Recall",    m["Recall"],     0.65, 0.40) +
            mbar("F1 Score",  m["F1"],         0.50, 0.25) +
            mbar("Precision", m["Precision"],  0.50, 0.20),
            unsafe_allow_html=True)

    st.markdown(f"""
    <div style="padding:14px 16px 16px;margin-top:4px;border-top:1px solid {C_BORDER};">
        <div style="font-size:9px;font-weight:800;color:{C_MUTED};
                    letter-spacing:0.14em;text-transform:uppercase;margin-bottom:10px;">
            AUTO SCHEDULER</div>
        <div style="background:{C_BG};border:1px solid {C_BORDER};border-radius:10px;
                    padding:12px 14px;font-size:11px;font-family:'JetBrains Mono',monospace;
                    line-height:2;color:{C_MUTED};">
            $ python scheduler.py<br>
            <span style="color:{C_ACCENT_L};">→ Scan inbox/ mỗi 60s</span>
        </div>
    </div>""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════
# TABS
# ═══════════════════════════════════════════════════════════
FLOW_STEPS = [
    ("📂", "UPLOAD & CLEAN"),
    ("🤖", "DỰ ĐOÁN"),
    ("📊", "TỔNG QUAN"),
    ("🔍", "PHÂN TÍCH"),
    ("🔎", "TRA CỨU"),
]

tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📂  UPLOAD & CLEAN",
    "🤖  DỰ ĐOÁN",
    "📊  TỔNG QUAN",
    "🔍  PHÂN TÍCH",
    "🔎  TRA CỨU",
])

# ═══════════════════════════════════════════════════════════
# TAB 1 — UPLOAD & CHUẨN HÓA
# ═══════════════════════════════════════════════════════════
with tab1:
    page_header("UPLOAD & CHUẨN HÓA DỮ LIỆU",
                "Bước 1: Upload CSV bất kỳ → tự động detect cột → làm sạch → sẵn sàng cho dự đoán")
    step_flow(FLOW_STEPS, 0)

    uploaded = st.file_uploader(
        "Kéo thả hoặc chọn file CSV giao dịch",
        type=["csv"],
        help="Hỗ trợ bất kỳ CSV nào — hệ thống sẽ tự động map cột",
        key="upload_clean_csv"
    )

    if uploaded is None:
        st.markdown(f"""
        <div style="background:{C_BG2};border:1px solid {C_BORDER};border-radius:16px;
                    padding:40px;text-align:center;margin-top:8px;">
            <div style="font-size:2.5rem;margin-bottom:16px;">📂</div>
            <div style="font-size:15px;font-weight:700;color:{C_TEXT};margin-bottom:8px;">
                Chưa có file nào được upload</div>
            <div style="font-size:13px;color:{C_MUTED};max-width:480px;margin:0 auto;line-height:1.7;">
                Upload file CSV giao dịch ngân hàng của bạn.<br>
                Hệ thống sẽ tự động phát hiện cột, map tên, xử lý missing values
                và chuẩn hóa trước khi đưa vào model dự đoán.
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Sample format hint
        sec_header("CẤU TRÚC CỘT KỲ VỌNG")
        hint_df = pd.DataFrame({
            "Tên cột chuẩn": REQUIRED_COLS,
            "Tên cột thay thế chấp nhận được": [
                "Transaction_Amount (in Million)", "Account_Balance (in Million)",
                "—", "—", "—",
                "Is_International_Transaction", "—",
                "Unusual_Time_Transaction", "—", "—", "—",
                "Avg_Transaction_Amount (in Million)",
                "Max_Transaction_Last_24h (in Million)",
                "—", "—", "—", "—",
            ][:len(REQUIRED_COLS)]
        })
        st.dataframe(hint_df, use_container_width=True, hide_index=True, height=300)

    else:
        df_raw = pd.read_csv(uploaded)
        st.success(f"✓ Đọc **{len(df_raw):,}** dòng · **{len(df_raw.columns)}** cột từ `{uploaded.name}`")

        # ── Bước 1: Detect & Map cột ─────────────────────────
        sec_header("BƯỚC 1 — PHÁT HIỆN & MAP CỘT")

        # Auto-rename known aliases
        df_mapped = df_raw.rename(columns=RENAME_MAP)
        existing   = set(df_mapped.columns)
        matched    = [c for c in REQUIRED_COLS if c in existing]
        missing    = [c for c in REQUIRED_COLS if c not in existing]
        extra      = [c for c in existing if c not in REQUIRED_COLS]

        c1, c2, c3 = st.columns(3)
        with c1: st.markdown(kpi(f"{len(matched)}/{len(REQUIRED_COLS)}", "CỘT KHỚP", "green"), unsafe_allow_html=True)
        with c2: st.markdown(kpi(str(len(missing)), "CỘT THIẾU", "red" if missing else "green"), unsafe_allow_html=True)
        with c3: st.markdown(kpi(str(len(extra)), "CỘT THÊM", "indigo"), unsafe_allow_html=True)

        st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

        # Hiển thị trạng thái từng cột
        col_l, col_r = st.columns(2)
        with col_l:
            st.markdown(f'<div style="font-size:11px;font-weight:700;color:{C_LOW};'
                        f'text-transform:uppercase;letter-spacing:0.1em;margin-bottom:8px;">'
                        f'✓ CỘT ĐÃ KHỚP ({len(matched)})</div>', unsafe_allow_html=True)
            for c in matched:
                st.markdown(f'<div class="col-row col-match">'
                            f'<span style="color:{C_LOW};font-size:13px;">✓</span>'
                            f'<span style="font-weight:600;color:{C_TEXT};">{c}</span>'
                            f'</div>', unsafe_allow_html=True)

        with col_r:
            if missing:
                st.markdown(f'<div style="font-size:11px;font-weight:700;color:{C_HIGH};'
                            f'text-transform:uppercase;letter-spacing:0.1em;margin-bottom:8px;">'
                            f'✗ CỘT THIẾU ({len(missing)}) — sẽ điền 0</div>', unsafe_allow_html=True)
                for c in missing:
                    # Cho phép map thủ công
                    user_map = st.selectbox(
                        f"{c}",
                        ["— Bỏ qua (điền 0) —"] + list(extra),
                        key=f"map_{c}",
                        label_visibility="visible"
                    )
                    if user_map != "— Bỏ qua (điền 0) —":
                        df_mapped = df_mapped.rename(columns={user_map: c})
                    else:
                        df_mapped[c] = 0
            else:
                st.markdown(f"""
                <div style="background:{C_LOW}10;border:1px solid {C_LOW}30;
                            border-radius:10px;padding:20px;text-align:center;">
                    <div style="font-size:1.5rem;margin-bottom:8px;">🎉</div>
                    <div style="color:{C_LOW};font-weight:700;font-size:13px;">
                        Tất cả cột đều khớp!</div>
                </div>""", unsafe_allow_html=True)

        # ── Bước 2: Missing values ────────────────────────────
        sec_header("BƯỚC 2 — XỬ LÝ MISSING VALUES")

        missing_stats = df_mapped.isnull().sum()
        missing_stats = missing_stats[missing_stats > 0]

        if len(missing_stats) == 0:
            st.markdown(f"""
            <div style="background:{C_LOW}10;border:1px solid {C_LOW}30;border-radius:10px;
                        padding:16px 20px;display:flex;align-items:center;gap:12px;">
                <span style="font-size:1.2rem;">✓</span>
                <span style="color:{C_LOW};font-weight:600;font-size:13px;">
                    Không có missing values — data sạch hoàn toàn!</span>
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown(f'<div style="color:{C_MUTED};font-size:13px;margin-bottom:12px;">'
                        f'Phát hiện <b style="color:{C_MED};">{missing_stats.sum():,}</b> '
                        f'ô trống trong <b>{len(missing_stats)}</b> cột.</div>',
                        unsafe_allow_html=True)

            miss_col1, miss_col2 = st.columns([2, 1])
            with miss_col1:
                miss_df = pd.DataFrame({
                    "Cột": missing_stats.index,
                    "Số ô trống": missing_stats.values,
                    "Tỷ lệ": [f"{v/len(df_mapped)*100:.1f}%" for v in missing_stats.values],
                })
                st.dataframe(miss_df, use_container_width=True, hide_index=True)
            with miss_col2:
                fill_method = st.radio(
                    "Cách điền missing:",
                    ["Median (số) / Mode (chữ) — khuyến nghị",
                     "Điền 0 cho tất cả",
                     "Xóa dòng thiếu"],
                    key="fill_method"
                )

        # ── Bước 3: Preview & Confirm ─────────────────────────
        sec_header("BƯỚC 3 — XEM TRƯỚC & XÁC NHẬN")

        # Tính toán preview stats
        n_rows_raw   = len(df_raw)
        n_missing_total = df_mapped.isnull().sum().sum()

        st.markdown(f"""
        <div style="display:flex;gap:12px;margin-bottom:16px;">
            <div style="flex:1;background:{C_BG2};border:1px solid {C_BORDER};
                        border-radius:12px;padding:16px 20px;">
                <div style="font-size:10px;color:{C_MUTED};font-weight:700;
                            text-transform:uppercase;letter-spacing:0.1em;margin-bottom:8px;">
                    TRƯỚC KHI CHUẨN HÓA</div>
                <div style="font-size:13px;color:{C_TEXT};line-height:2;">
                    📄 {n_rows_raw:,} dòng · {len(df_raw.columns)} cột<br>
                    ⚠ {n_missing_total:,} ô trống<br>
                    📋 {len(missing)} cột thiếu so với model
                </div>
            </div>
            <div style="display:flex;align-items:center;font-size:20px;color:{C_ACCENT};">→</div>
            <div style="flex:1;background:{C_ACCENT}08;border:1px solid {C_ACCENT}40;
                        border-radius:12px;padding:16px 20px;">
                <div style="font-size:10px;color:{C_ACCENT_L};font-weight:700;
                            text-transform:uppercase;letter-spacing:0.1em;margin-bottom:8px;">
                    SAU KHI CHUẨN HÓA</div>
                <div style="font-size:13px;color:{C_TEXT};line-height:2;">
                    ✓ {n_rows_raw:,} dòng · {len(REQUIRED_COLS)}+ cột chuẩn<br>
                    ✓ 0 ô trống (đã điền tự động)<br>
                    ✓ Sẵn sàng đưa vào model ML
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Preview 5 dòng đầu
        st.markdown(f'<div style="font-size:11px;color:{C_MUTED};font-weight:600;'
                    f'margin-bottom:8px;">XEM TRƯỚC 5 DÒNG ĐẦU (data gốc):</div>',
                    unsafe_allow_html=True)
        st.dataframe(df_raw.head(), use_container_width=True, hide_index=True)

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        # Nút confirm
        if st.button("✓  XÁC NHẬN & LÀM SẠCH DỮ LIỆU", type="primary", use_container_width=True):
            with st.spinner("Đang chuẩn hóa dữ liệu…"):
                # Xác định fill method
                fm = st.session_state.get("fill_method", "")
                if "Xóa" in fm:
                    df_mapped = df_mapped.dropna()
                elif "Điền 0" in fm:
                    df_mapped = df_mapped.fillna(0)
                # else: clean_uploaded_df tự xử lý median/mode

                df_clean = clean_uploaded_df(df_mapped, {})
                st.session_state["clean_df"]      = df_clean
                st.session_state["source_filename"] = uploaded.name
                # Reset predict khi upload file mới
                if "predict_result" in st.session_state:
                    del st.session_state["predict_result"]

            st.success(f"✓ Đã chuẩn hóa **{len(df_clean):,}** dòng — chuyển sang **Tab 🤖 DỰ ĐOÁN**")
            st.balloons()


# ═══════════════════════════════════════════════════════════
# TAB 2 — DỰ ĐOÁN
# ═══════════════════════════════════════════════════════════
with tab2:
    page_header("DỰ ĐOÁN FRAUD", "Bước 2: Chạy ML pipeline → phân loại rủi ro từng giao dịch")
    step_flow(FLOW_STEPS, 1)

    if "clean_df" not in st.session_state:
        st.markdown(f"""
        <div style="background:{C_BG2};border:1px dashed {C_BORDER};border-radius:16px;
                    padding:48px;text-align:center;">
            <div style="font-size:2.5rem;margin-bottom:16px;">⬅️</div>
            <div style="font-size:15px;font-weight:700;color:{C_TEXT};margin-bottom:8px;">
                Chưa có dữ liệu</div>
            <div style="font-size:13px;color:{C_MUTED};">
                Quay lại <strong>Tab 📂 UPLOAD & CLEAN</strong> để upload và chuẩn hóa dữ liệu trước
            </div>
        </div>""", unsafe_allow_html=True)
    elif not os.path.exists(MODEL_PATH):
        st.warning("⚠️ Model chưa được train. Nhấn **▶ RUN FULL PIPELINE** ở sidebar trước.")
    else:
        df_clean = st.session_state["clean_df"]
        fname    = st.session_state.get("source_filename", "data.csv")

        st.markdown(f"""
        <div style="background:{C_LOW}10;border:1px solid {C_LOW}30;border-radius:12px;
                    padding:16px 20px;display:flex;align-items:center;gap:16px;margin-bottom:20px;">
            <span style="font-size:1.5rem;">✓</span>
            <div>
                <div style="font-weight:700;color:{C_LOW};font-size:13px;">Dữ liệu đã sẵn sàng</div>
                <div style="color:{C_MUTED};font-size:12px;margin-top:2px;">
                    {len(df_clean):,} giao dịch từ <code>{fname}</code> — đã qua chuẩn hóa
                </div>
            </div>
        </div>""", unsafe_allow_html=True)

        if st.button("🚀  CHẠY DỰ ĐOÁN FRAUD", type="primary", use_container_width=True):
            with st.spinner("Đang tạo features và dự đoán…"):
                try:
                    result = predict_from_clean(df_clean)
                    st.session_state["predict_result"] = result
                    st.success(f"✓ Dự đoán hoàn thành — {len(result):,} giao dịch đã được phân loại!")
                except Exception as e:
                    st.error(f"Lỗi: {e}")

        if "predict_result" in st.session_state:
            result = st.session_state["predict_result"]
            rc = risk_counts_from_level(result)
            n = len(result)
            n_high = rc["HIGH"]
            n_med = rc["MEDIUM"]
            avg_prob = result["Fraud_Probability"].mean() * 100 if "Fraud_Probability" in result.columns else 0

            st.markdown(f"""
            <div style="background:{C_ACCENT}08;border:1px solid {C_ACCENT}35;border-radius:14px;
                        padding:18px 22px;margin:16px 0 18px 0;">
                <div style="font-size:13px;color:{C_TEXT};font-weight:800;margin-bottom:8px;">
                    Dự đoán hoàn tất</div>
                <div style="font-size:12px;color:{C_MUTED};line-height:1.7;">
                    Tab này tập trung vào danh sách ưu tiên để kiểm tra và xuất báo cáo. Các biểu đồ tổng quan được đặt ở tab <b>Tổng quan</b>.
                </div>
            </div>""", unsafe_allow_html=True)

            a,b,c = st.columns(3)
            with a: insight_card("HIGH RISK", f"{n_high:,}", "Giao dịch cần ưu tiên rà soát ngay", C_HIGH)
            with b: insight_card("MEDIUM RISK", f"{n_med:,}", "Giao dịch đưa vào queue giám sát", C_MED)
            with c: insight_card("AVG FRAUD SCORE", f"{avg_prob:.1f}%", f"Trung bình trên {n:,} giao dịch", C_ACCENT_L)

            sec_header("KẾT QUẢ DỰ ĐOÁN — DANH SÁCH ƯU TIÊN")
            view_mode = st.radio(
                "Chế độ xem",
                ["Chỉ giao dịch cần xử lý", "Tóm tắt theo nhóm", "Bảng đầy đủ rút gọn"],
                horizontal=True,
                key="pred_view_mode"
            )

            compact_cols = [c for c in [
                "Transaction_ID", "Fraud_Probability", "Risk_Level", "Transaction_Amount",
                "Transaction_Type", "Merchant_Category", "Is_International", "Unusual_Time"
            ] if c in result.columns]

            if view_mode == "Chỉ giao dịch cần xử lý":
                priority = result[result["Risk_Level"].isin(["HIGH", "MEDIUM"])].sort_values("Fraud_Probability", ascending=False)
                st.caption("Hiển thị HIGH/MEDIUM trước để người vận hành không phải đọc toàn bộ bảng lớn.")
                show_n = st.slider("Số dòng hiển thị", 25, 500, 100, step=25, key="pred_priority_n")
                st.dataframe(
                    priority[compact_cols].head(show_n).style.applymap(style_risk_cell, subset=["Risk_Level"]),
                    use_container_width=True, height=420, hide_index=True
                )
            elif view_mode == "Tóm tắt theo nhóm":
                group_col = st.selectbox("Nhóm theo", [c for c in ["Merchant_Category", "Transaction_Type", "Card_Type", "Is_International", "Unusual_Time"] if c in result.columns], key="pred_group_col")
                g = group_risk_summary(result, group_col)
                g_show = g[[group_col, "Total", "High", "Medium", "High_Rate", "Avg_Prob_%"]].rename(columns={
                    group_col: "Nhóm", "Total": "Tổng GD", "High": "HIGH", "Medium": "MEDIUM",
                    "High_Rate": "Tỷ lệ HIGH (%)", "Avg_Prob_%": "Xác suất TB (%)"
                }).round(2)
                st.dataframe(
                    g_show.style.background_gradient(subset=["Tỷ lệ HIGH (%)", "Xác suất TB (%)"], cmap="Reds"),
                    use_container_width=True, hide_index=True, height=360
                )
            else:
                st.caption("Bảng đã rút bớt các trường kỹ thuật để dễ đọc; file Excel vẫn giữ đầy đủ cột.")
                st.dataframe(
                    result[compact_cols].sort_values("Fraud_Probability", ascending=False).head(1000).style.applymap(style_risk_cell, subset=["Risk_Level"]),
                    use_container_width=True, height=420, hide_index=True
                )

            col_dl1, col_dl2 = st.columns(2)
            with col_dl1:
                eb = df_to_excel_bytes(result)
                st.download_button("📥 XUẤT EXCEL ĐẦY ĐỦ", data=eb,
                    file_name=f"fraud_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            with col_dl2:
                if st.button("🌐 XUẤT BÁO CÁO HTML", use_container_width=True):
                    try:
                        from email_alert import _build_html
                        df_high = result[result["Risk_Level"]=="HIGH"].sort_values("Fraud_Probability", ascending=False)
                        html_content = _build_html(df_high, result, fname)
                        os.makedirs(REPORTS_DIR, exist_ok=True)
                        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                        html_path = os.path.join(REPORTS_DIR, f"fraud_report_{ts}.html")
                        with open(html_path, "w", encoding="utf-8") as f:
                            f.write(html_content)
                        st.success(f"✓ Đã lưu → output/reports/fraud_report_{ts}.html")
                        st.download_button("📥 Download HTML", data=html_content.encode("utf-8"),
                            file_name=f"fraud_report_{ts}.html", mime="text/html")
                    except Exception as e:
                        st.error(f"Lỗi: {e}")

            fi_path = os.path.join(OUTPUT_DIR, "feature_importance.png")
            if os.path.exists(fi_path):
                sec_header("FEATURE IMPORTANCE — TOP 10")
                st.markdown(f"""
                <div style="background:{C_BG2};border:1px solid {C_BORDER};border-radius:12px;padding:16px 20px;margin-bottom:12px;">
                    <div style="color:{C_TEXT};font-weight:800;font-size:13px;margin-bottom:8px;">Feature Importance dùng để làm gì?</div>
                    <div style="color:{C_MUTED};font-size:12px;line-height:1.8;">
                        Đây là bảng xếp hạng 10 biến đầu vào ảnh hưởng mạnh nhất đến quyết định của model.
                        Biểu đồ này không nói “giao dịch A chắc chắn gian lận vì biến X”, mà cho biết khi model học toàn bộ dữ liệu,
                        những biến nào thường giúp model phân biệt rủi ro tốt nhất. Dùng nó để giải thích model, chọn biến cần theo dõi,
                        và kiểm tra xem model có đang phụ thuộc vào các tín hiệu hợp lý như lịch sử fraud, thời gian bất thường,
                        giao dịch quốc tế, khoảng cách, số lần giao dịch lỗi… hay không.
                    </div>
                </div>""", unsafe_allow_html=True)
                st.image(fi_path, use_column_width=True)


# ═══════════════════════════════════════════════════════════
# TAB 3 — TỔNG QUAN (dựa trên kết quả dự đoán từ Tab 2)
# ═══════════════════════════════════════════════════════════
with tab3:
    page_header("TỔNG QUAN DỰ ĐOÁN", "Bước 3: Phân tích kết quả ML — Risk Level từ model")
    step_flow(FLOW_STEPS, 2)

    # Yêu cầu phải có predict_result từ Tab 2
    if "predict_result" not in st.session_state:
        st.markdown(f"""
        <div style="background:{C_BG2};border:1px dashed {C_BORDER};border-radius:16px;
                    padding:48px;text-align:center;">
            <div style="font-size:2.5rem;margin-bottom:16px;">⬅️</div>
            <div style="font-size:15px;font-weight:700;color:{C_TEXT};margin-bottom:8px;">
                Chưa có kết quả dự đoán</div>
            <div style="font-size:13px;color:{C_MUTED};line-height:1.8;">
                Vui lòng thực hiện theo thứ tự:<br>
                <strong>Tab 📂 UPLOAD & CLEAN</strong> → upload CSV<br>
                <strong>Tab 🤖 DỰ ĐOÁN</strong> → chạy dự đoán<br>
                Sau đó quay lại tab này để xem phân tích.
            </div>
        </div>""", unsafe_allow_html=True)
    else:
        df   = st.session_state["predict_result"]
        fname = st.session_state.get("source_filename", "file đã upload")
        st.markdown(f'<div style="font-size:11px;color:{C_MUTED};margin-bottom:16px;">📂 {fname} · {len(df):,} giao dịch · kết quả từ model ML</div>', unsafe_allow_html=True)

        rc = df["Risk_Level"].value_counts(); n = len(df)
        n_high, n_med, n_low = rc.get("HIGH",0), rc.get("MEDIUM",0), rc.get("LOW",0)
        p_h, p_m, p_l = n_high/n*100, n_med/n*100, n_low/n*100
        avg_prob = df["Fraud_Probability"].mean()*100

        c1, c2, c3, c4 = st.columns(4)
        with c1: st.markdown(kpi(f"{n:,}", "TỔNG GIAO DỊCH"), unsafe_allow_html=True)
        with c2: st.markdown(kpi(f"{n_high:,}", "HIGH RISK", "red"), unsafe_allow_html=True)
        with c3: st.markdown(kpi(f"{n_med:,}", "MEDIUM RISK", "amber"), unsafe_allow_html=True)
        with c4: st.markdown(kpi(f"{avg_prob:.1f}%", "AVG PROB", "indigo"), unsafe_allow_html=True)

        st.markdown(f"""
        <div style="background:{C_BG2};border:1px solid {C_BORDER};border-radius:12px;padding:20px 24px;margin:16px 0 24px 0;">
            <div style="font-size:10px;color:{C_MUTED};font-weight:700;text-transform:uppercase;letter-spacing:0.12em;margin-bottom:12px;">PHÂN BỔ RỦI RO TỔNG THỂ</div>
            <div style="display:flex;height:14px;border-radius:7px;overflow:hidden;gap:2px;">
                <div style="width:{p_h:.1f}%;background:{C_HIGH};border-radius:5px 0 0 5px;"></div>
                <div style="width:{p_m:.1f}%;background:{C_MED};"></div>
                <div style="width:{p_l:.1f}%;background:{C_LOW};border-radius:0 5px 5px 0;"></div>
            </div>
            <div style="display:flex;gap:24px;margin-top:10px;font-size:12px;font-weight:600;">
                <span style="color:{C_HIGH};">● HIGH {p_h:.1f}% ({n_high:,})</span>
                <span style="color:{C_MED};">● MEDIUM {p_m:.1f}% ({n_med:,})</span>
                <span style="color:{C_LOW};">● LOW {p_l:.1f}% ({n_low:,})</span>
            </div>
        </div>""", unsafe_allow_html=True)

        high_df = df[df["Risk_Level"] == "HIGH"]
        col_a, col_b = st.columns([3, 2])
        with col_a:
            sec_header("RISK LEVEL THEO MERCHANT CATEGORY")
            if "Merchant_Category" in df.columns:
                pivot = df.groupby(["Merchant_Category","Risk_Level"]).size().unstack(fill_value=0)
                cats = pivot.index.tolist(); x=np.arange(len(cats)); w=.25
                fig, ax = chart_fig(8,4.8)
                for i,(lv,clr) in enumerate([("HIGH",C_HIGH),("MEDIUM",C_MED),("LOW",C_LOW)]):
                    ax.bar(x+i*w, pivot[lv] if lv in pivot.columns else 0, w, label=lv, color=clr, alpha=.85, zorder=3)
                ax.set_xticks(x+w); ax.set_xticklabels(cats, rotation=22, ha="right", color=C_MUTED, fontsize=9)
                ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v,_:f"{v:,.0f}")); ax.legend(facecolor=C_BG2, labelcolor=C_TEXT, fontsize=9)
                ax.set_title("Merchant nào tạo nhiều rủi ro nhất?", fontsize=11, color=C_TEXT, fontweight="700")
                fig.tight_layout(); st.pyplot(fig); plt.close()
        with col_b:
            sec_header("FRAUD PROBABILITY — PHÂN PHỐI")
            fig4, ax4 = chart_fig(5,4.8)
            for lv,clr in [("LOW",C_LOW),("MEDIUM",C_MED),("HIGH",C_HIGH)]:
                vals=df.loc[df["Risk_Level"]==lv,"Fraud_Probability"].dropna()
                if len(vals): ax4.hist(vals, bins=30, color=clr, alpha=.65, label=lv, density=True)
            ax4.axvline(.30, color=C_MED, linestyle="--", linewidth=1.3); ax4.axvline(.60, color=C_HIGH, linestyle="--", linewidth=1.3)
            ax4.set_xlabel("Fraud Probability", color=C_MUTED); ax4.legend(facecolor=C_BG2, labelcolor=C_TEXT, fontsize=9)
            ax4.set_title("Ngưỡng LOW / MEDIUM / HIGH", fontsize=11, color=C_TEXT, fontweight="700")
            fig4.tight_layout(); st.pyplot(fig4); plt.close()

        col_c, col_d = st.columns([3, 2])
        with col_c:
            sec_header("HEATMAP — HIGH RISK THEO GIỜ × NGÀY")
            if "Hour" in high_df.columns and "DayOfWeek" in high_df.columns and len(high_df):
                piv = high_df.dropna(subset=["Hour","DayOfWeek"]).copy(); piv["Hour"]=piv["Hour"].astype(int); piv["DayOfWeek"]=piv["DayOfWeek"].astype(int)
                heat = piv.groupby(["DayOfWeek","Hour"]).size().unstack(fill_value=0).reindex(index=range(7), columns=range(24), fill_value=0)
                fig3, ax3 = chart_fig(9,4)
                im=ax3.imshow(heat.values, aspect="auto", cmap="RdPu", interpolation="nearest")
                ax3.set_yticks(range(7)); ax3.set_yticklabels(["Mon","Tue","Wed","Thu","Fri","Sat","Sun"], color=C_MUTED, fontsize=9)
                ax3.set_xticks(range(0,24,2)); ax3.set_xticklabels([f"{h:02d}h" for h in range(0,24,2)], rotation=45, ha="right", color=C_MUTED, fontsize=8)
                cb=plt.colorbar(im, ax=ax3); cb.ax.tick_params(colors=C_MUTED); cb.set_label("Số HIGH Risk", color=C_MUTED)
                ax3.grid(False); fig3.tight_layout(); st.pyplot(fig3); plt.close()
        with col_d:
            sec_header("TRANSACTION TYPE — TỶ LỆ HIGH")
            if "Transaction_Type" in df.columns:
                g=group_risk_summary(df,"Transaction_Type")
                fig5, ax5 = chart_fig(5,4)
                bars=ax5.barh(g["Transaction_Type"], g["High_Rate"], color=C_HIGH, alpha=.85)
                ax5.invert_yaxis(); ax5.set_xlabel("% HIGH", color=C_MUTED); ax5.set_title("Tỷ lệ HIGH theo loại GD", color=C_TEXT, fontsize=11, fontweight="700")
                fig5.tight_layout(); st.pyplot(fig5); plt.close()

        sec_header("TỔ HỢP RỦI RO — INTERNATIONAL × UNUSUAL TIME")
        if "Is_International" in df.columns and "Unusual_Time" in df.columns:
            df_c=df.copy(); df_c["Tổ hợp"]=df_c["Is_International"].astype(str)+" / "+df_c["Unusual_Time"].astype(str)
            cs=df_c.groupby("Tổ hợp").agg(**{"Tổng GD":("Risk_Level","count"),"HIGH Risk":("Risk_Level",lambda x:(x=="HIGH").sum()),"MEDIUM Risk":("Risk_Level",lambda x:(x=="MEDIUM").sum()),"Avg Prob (%)":("Fraud_Probability",lambda x:x.mean()*100)}).reset_index()
            cs["High Rate (%)"]=(cs["HIGH Risk"]/cs["Tổng GD"]*100).round(2); cs["Avg Prob (%)"]=cs["Avg Prob (%)"].round(2)
            cs=cs.sort_values("High Rate (%)", ascending=False)
            st.dataframe(cs.style.background_gradient(subset=["High Rate (%)","Avg Prob (%)"], cmap="Reds").bar(subset=["HIGH Risk","MEDIUM Risk"], color=C_ACCENT), use_container_width=True, hide_index=True, height=220)


# ═══════════════════════════════════════════════════════════
# TAB 4 — PHÂN TÍCH RỦI RO (dựa trên kết quả dự đoán)
# ═══════════════════════════════════════════════════════════
with tab4:
    page_header("PHÂN TÍCH RỦI RO", "Bước 4: Lọc sâu và khám phá pattern gian lận từ kết quả ML")
    step_flow(FLOW_STEPS, 3)

    if "predict_result" not in st.session_state:
        st.markdown(f"""
        <div style="background:{C_BG2};border:1px dashed {C_BORDER};border-radius:16px;
                    padding:48px;text-align:center;">
            <div style="font-size:2.5rem;margin-bottom:16px;">⬅️</div>
            <div style="font-size:15px;font-weight:700;color:{C_TEXT};margin-bottom:8px;">
                Chưa có kết quả dự đoán</div>
            <div style="font-size:13px;color:{C_MUTED};line-height:1.8;">
                Vui lòng thực hiện theo thứ tự:<br>
                <strong>Tab 📂 UPLOAD & CLEAN</strong> → upload CSV<br>
                <strong>Tab 🤖 DỰ ĐOÁN</strong> → chạy dự đoán<br>
                Sau đó quay lại tab này để phân tích.
            </div>
        </div>""", unsafe_allow_html=True)
    else:
        df4 = st.session_state["predict_result"]
        sec_header("1) BỘ LỌC PHÂN TÍCH")
        f1, f2, f3, f4 = st.columns(4)
        with f1: sel_risk = st.selectbox("RISK LEVEL", ["Tất cả","HIGH","MEDIUM","LOW"], key="t4_risk")
        with f2:
            opts_tx = ["Tất cả"] + (sorted(df4["Transaction_Type"].dropna().unique().tolist()) if "Transaction_Type" in df4.columns else [])
            sel_tx = st.selectbox("TRANSACTION TYPE", opts_tx, key="t4_tx")
        with f3:
            opts_mc = ["Tất cả"] + (sorted(df4["Merchant_Category"].dropna().unique().tolist()) if "Merchant_Category" in df4.columns else [])
            sel_mc = st.selectbox("MERCHANT", opts_mc, key="t4_mc")
        with f4: sel_intl = st.selectbox("INTERNATIONAL", ["Tất cả","Yes","No"], key="t4_intl")

        dff = df4.copy()
        if sel_risk != "Tất cả": dff = dff[dff["Risk_Level"] == sel_risk]
        if sel_tx != "Tất cả" and "Transaction_Type" in dff.columns: dff = dff[dff["Transaction_Type"] == sel_tx]
        if sel_mc != "Tất cả" and "Merchant_Category" in dff.columns: dff = dff[dff["Merchant_Category"] == sel_mc]
        if sel_intl != "Tất cả" and "Is_International" in dff.columns: dff = dff[dff["Is_International"] == sel_intl]

        n_filt=len(dff); rc_f=dff["Risk_Level"].value_counts(); n_h=rc_f.get("HIGH",0); avg_p=dff["Fraud_Probability"].mean()*100 if n_filt else 0
        c1,c2,c3,c4=st.columns(4)
        with c1: insight_card("Tập đang xem", f"{n_filt:,}", "Số giao dịch sau khi áp dụng bộ lọc")
        with c2: insight_card("HIGH risk", f"{n_h:,}", f"{(n_h/max(n_filt,1)*100):.1f}% trong tập đang xem", C_HIGH)
        with c3: insight_card("Avg fraud prob", f"{avg_p:.1f}%", "Xác suất trung bình của tập lọc", C_MED if avg_p>=30 else C_LOW)
        with c4:
            top_seg = "—"
            if "Merchant_Category" in dff.columns and len(dff):
                top_seg = str(group_risk_summary(dff,"Merchant_Category").iloc[0]["Merchant_Category"])
            insight_card("Nhóm rủi ro nhất", top_seg, "Dựa trên High Rate trong tập lọc", C_ACCENT_L)

        sec_header("2) DRIVER CHÍNH — NHÓM NÀO LÀM RỦI RO TĂNG?")
        cols = [c for c in ["Merchant_Category","Transaction_Type","Card_Type","Is_International","Unusual_Time"] if c in dff.columns]
        if cols and len(dff):
            pick = st.selectbox("Chọn chiều phân tích", cols, key="driver_dimension")
            g = group_risk_summary(dff, pick).head(12)
            colx, coly = st.columns([3,2])
            with colx:
                fig, ax = chart_fig(8,4.5)
                y = np.arange(len(g))
                ax.barh(y, g["High_Rate"], color=C_HIGH, alpha=.85, label="High Rate")
                ax.set_yticks(y); ax.set_yticklabels(g[pick].astype(str), color=C_MUTED, fontsize=9)
                ax.invert_yaxis(); ax.set_xlabel("% HIGH risk", color=C_MUTED)
                ax.set_title(f"Tỷ lệ HIGH theo {pick}", color=C_TEXT, fontsize=11, fontweight="700")
                fig.tight_layout(); st.pyplot(fig); plt.close()
            with coly:
                st.dataframe(g[[pick,"Total","High","Medium","High_Rate","Avg_Prob_%"]].round(2).style.background_gradient(subset=["High_Rate","Avg_Prob_%"], cmap="Reds"), use_container_width=True, hide_index=True, height=360)

        sec_header("3) LIÊN HỆ GIỮA SỐ TIỀN, SỐ DƯ VÀ XÁC SUẤT FRAUD")
        cs1, cs2 = st.columns(2)
        with cs1:
            if {"Transaction_Amount","Account_Balance","Risk_Level"}.issubset(dff.columns) and len(dff):
                smp=dff.dropna(subset=["Transaction_Amount","Account_Balance"]).sample(min(2500,len(dff)), random_state=42)
                fig7, ax7=chart_fig(7,4.5)
                for level,clr,al,sz in [("LOW",C_LOW,.25,8),("MEDIUM",C_MED,.55,12),("HIGH",C_HIGH,.9,16)]:
                    s2=smp[smp["Risk_Level"]==level]
                    ax7.scatter(s2["Transaction_Amount"], s2["Account_Balance"], c=clr, s=sz, alpha=al, label=level, edgecolors="none")
                ax7.legend(facecolor=C_BG2, labelcolor=C_TEXT, fontsize=9); ax7.set_xlabel("Transaction Amount", color=C_MUTED); ax7.set_ylabel("Account Balance", color=C_MUTED)
                fig7.tight_layout(); st.pyplot(fig7); plt.close()
        with cs2:
            if {"Transaction_Amount","Risk_Level"}.issubset(dff.columns) and len(dff):
                fig8, ax8=chart_fig(6,4.5)
                data_box=[dff[dff["Risk_Level"]==lv]["Transaction_Amount"].dropna().values for lv in ["LOW","MEDIUM","HIGH"]]
                bp=ax8.boxplot(data_box, patch_artist=True, medianprops={"color":"white","linewidth":2.2})
                for box,clr in zip(bp["boxes"],[C_LOW,C_MED,C_HIGH]): box.set_facecolor(clr+"33"); box.set_edgecolor(clr)
                ax8.set_xticklabels(["LOW","MEDIUM","HIGH"], color=C_TEXT, fontweight="700"); ax8.set_title("Amount theo Risk Level", color=C_TEXT, fontsize=11, fontweight="700")
                fig8.tight_layout(); st.pyplot(fig8); plt.close()

        sec_header("4) QUEUE XỬ LÝ — TOP GIAO DỊCH RỦI RO")
        show_cols=[c for c in ["Transaction_ID","Fraud_Probability","Risk_Level","Transaction_Amount","Transaction_Type","Merchant_Category","Card_Type","Is_International","Unusual_Time","Account_Balance"] if c in dff.columns]
        queue=dff.sort_values("Fraud_Probability", ascending=False).head(300)
        st.dataframe(queue[show_cols].style.applymap(style_risk_cell, subset=["Risk_Level"]), use_container_width=True, height=380, hide_index=True)


# ═══════════════════════════════════════════════════════════
# TAB 5 — TRA CỨU
# ═══════════════════════════════════════════════════════════
with tab5:
    page_header("TRA CỨU GIAO DỊCH", "Bước 5: Lọc và xem top rủi ro theo từng nhóm")
    step_flow(FLOW_STEPS, 4)

    if "predict_result" not in st.session_state:
        st.markdown(f"""
        <div style="background:{C_BG2};border:1px dashed {C_BORDER};border-radius:16px;
                    padding:48px;text-align:center;">
            <div style="font-size:2.5rem;margin-bottom:16px;">🔎</div>
            <div style="font-size:15px;font-weight:700;color:{C_TEXT};margin-bottom:8px;">
                Chưa có kết quả dự đoán</div>
            <div style="font-size:13px;color:{C_MUTED};">
                Vào <strong>Tab 🤖 DỰ ĐOÁN</strong> → chạy dự đoán trước
            </div>
        </div>""", unsafe_allow_html=True)
    else:
        result = st.session_state["predict_result"]

        col_s1,col_s2,col_s3,col_s4 = st.columns([2,2,2,1])
        with col_s1:
            filter_risk = st.selectbox("RISK LEVEL", ["Tất cả","HIGH","MEDIUM","LOW"], key="search_risk_level")
        with col_s2:
            tx_opts = ["Tất cả"] + (sorted(result["Transaction_Type"].dropna().unique().tolist())
                                    if "Transaction_Type" in result.columns else [])
            filter_tx = st.selectbox("TRANSACTION TYPE", tx_opts, key="search_transaction_type")
        with col_s3:
            mc_opts = ["Tất cả"] + (sorted(result["Merchant_Category"].dropna().unique().tolist())
                                    if "Merchant_Category" in result.columns else [])
            filter_mc = st.selectbox("MERCHANT", mc_opts, key="search_merchant")
        with col_s4:
            filter_intl = st.selectbox("INTL", ["Tất cả","Yes","No"], key="search_international")

        filtered = result.copy()
        if filter_risk != "Tất cả" and "Risk_Level"        in filtered.columns: filtered = filtered[filtered["Risk_Level"]        == filter_risk]
        if filter_tx   != "Tất cả" and "Transaction_Type"  in filtered.columns: filtered = filtered[filtered["Transaction_Type"]  == filter_tx]
        if filter_mc   != "Tất cả" and "Merchant_Category" in filtered.columns: filtered = filtered[filtered["Merchant_Category"] == filter_mc]
        if filter_intl != "Tất cả" and "Is_International"  in filtered.columns: filtered = filtered[filtered["Is_International"]  == filter_intl]

        n_found = len(filtered)
        n_high  = (filtered["Risk_Level"]=="HIGH").sum()   if "Risk_Level" in filtered.columns else 0
        n_med   = (filtered["Risk_Level"]=="MEDIUM").sum() if "Risk_Level" in filtered.columns else 0
        n_low   = (filtered["Risk_Level"]=="LOW").sum()    if "Risk_Level" in filtered.columns else 0

        c1,c2,c3,c4 = st.columns(4)
        with c1: st.markdown(kpi(f"{n_found:,}", "TÌM THẤY", "indigo"), unsafe_allow_html=True)
        with c2: st.markdown(kpi(f"{n_high:,}",  "HIGH RISK", "red"),    unsafe_allow_html=True)
        with c3: st.markdown(kpi(f"{n_med:,}",   "MEDIUM RISK", "amber"),unsafe_allow_html=True)
        with c4: st.markdown(kpi(f"{n_low:,}",   "LOW RISK", "green"),   unsafe_allow_html=True)

        if n_found == 0:
            st.markdown(f"""
            <div style="background:{C_BG2};border:1px dashed {C_BORDER};border-radius:10px;
                        padding:32px;text-align:center;color:{C_MUTED};margin-top:16px;">
                Không tìm thấy giao dịch nào khớp với điều kiện lọc.
            </div>""", unsafe_allow_html=True)
        else:
            sec_header("TOP RỦI RO THEO TỪNG NHÓM")
            cat_cols = [c for c in ["Transaction_Type","Merchant_Category","Card_Type"]
                        if c in filtered.columns]
            for cat_col in cat_cols:
                st.markdown(
                    f'<div style="color:{C_TEXT};font-weight:800;font-size:12px;'
                    f'text-transform:uppercase;letter-spacing:0.1em;'
                    f'margin:20px 0 10px 0;border-left:3px solid {C_ACCENT};'
                    f'padding-left:10px;">{cat_col.replace("_"," ")}</div>',
                    unsafe_allow_html=True)
                groups = filtered.groupby(cat_col).agg(
                    Total    =("Risk_Level","count"),
                    High     =("Risk_Level", lambda x: (x=="HIGH").sum()),
                    Avg_Prob =("Fraud_Probability","mean"),
                ).reset_index()
                groups["High_Rate"] = (groups["High"]/groups["Total"]*100).round(1)
                groups = groups.sort_values("High", ascending=False)
                max_h = max(groups["High"].max(), 1)

                for _, rg in groups.iterrows():
                    bar_w   = rg["High"]/max_h*100
                    prob_p  = rg["Avg_Prob"]*100
                    pc      = C_HIGH if prob_p>=60 else (C_MED if prob_p>=30 else C_LOW)
                    st.markdown(f"""
                    <div style="background:{C_BG2};border:1px solid {C_BORDER};
                                border-radius:10px;padding:14px 18px;margin-bottom:6px;
                                display:flex;align-items:center;gap:16px;">
                        <div style="width:140px;flex-shrink:0;">
                            <div style="font-size:13px;font-weight:700;color:{C_TEXT};">
                                {rg[cat_col]}</div>
                            <div style="font-size:11px;color:{C_MUTED};margin-top:2px;">
                                {rg['Total']:,} giao dịch</div>
                        </div>
                        <div style="flex:1;">
                            <div style="background:{C_BG3};border-radius:100px;
                                        height:8px;overflow:hidden;">
                                <div style="height:100%;width:{bar_w:.1f}%;
                                            background:linear-gradient(90deg,{C_HIGH},{C_MED});
                                            border-radius:100px;"></div>
                            </div>
                            <div style="font-size:11px;color:{C_MUTED};margin-top:5px;">
                                {rg['High']:,} HIGH risk · {rg['High_Rate']:.1f}% fraud rate
                            </div>
                        </div>
                        <div style="text-align:center;flex-shrink:0;width:72px;">
                            <div style="font-size:17px;font-weight:700;
                                        font-family:'JetBrains Mono',monospace;
                                        color:{pc};">{prob_p:.1f}%</div>
                            <div style="font-size:9px;color:{C_MUTED};font-weight:600;
                                        text-transform:uppercase;letter-spacing:0.08em;">
                                avg prob</div>
                        </div>
                    </div>""", unsafe_allow_html=True)

            # Chi tiết 1 giao dịch
            sec_header("XEM CHI TIẾT 1 GIAO DỊCH")
            pool = filtered[filtered["Risk_Level"]=="HIGH"].sort_values(
                "Fraud_Probability", ascending=False) if n_high>0 else filtered

            if "Transaction_ID" in pool.columns:
                pool_show = pool.sort_values("Fraud_Probability", ascending=False).head(300).copy()
                option_map = {risk_priority_label(r): str(r["Transaction_ID"]) for _, r in pool_show.iterrows()}
                sel_label = st.selectbox(
                    "CHỌN GIAO DỊCH — hiển thị ID + xác suất + loại + merchant + số tiền",
                    list(option_map.keys()), key="detail_select"
                )
                sel_id = option_map[sel_label]
                row = pool[pool["Transaction_ID"].astype(str)==sel_id].iloc[0]
            else:
                pool_show = pool.sort_values("Fraud_Probability", ascending=False).head(300).copy()
                option_map = {risk_priority_label(r): i for i, r in pool_show.iterrows()}
                sel_label = st.selectbox("CHỌN GIAO DỊCH", list(option_map.keys()), key="detail_select_noid")
                row = pool.loc[option_map[sel_label]]

            prob      = float(row.get("Fraud_Probability", 0))
            risk      = row.get("Risk_Level", "LOW")
            risk_color = {"HIGH":C_HIGH,"MEDIUM":C_MED,"LOW":C_LOW}.get(risk, C_MUTED)

            col_g, col_d = st.columns([1, 2])
            with col_g:
                pct   = prob*100
                angle = pct/100*180
                r,cx,cy = 70,90,90
                end_x = cx + r*math.cos(math.radians(180-angle))
                end_y = cy - r*math.sin(math.radians(180-angle))
                large = 1 if angle>180 else 0
                st.markdown(f"""
                <div style="text-align:center;background:{C_BG2};border:1px solid {C_BORDER};
                            border-radius:14px;padding:24px 20px;">
                    <svg width="180" height="110" viewBox="0 0 180 110">
                        <path d="M 20 90 A 70 70 0 0 1 160 90"
                              fill="none" stroke="{C_BG3}" stroke-width="16" stroke-linecap="round"/>
                        <path d="M 20 90 A 70 70 0 {large} 1 {end_x:.2f} {end_y:.2f}"
                              fill="none" stroke="{risk_color}" stroke-width="16"
                              stroke-linecap="round"/>
                        <text x="90" y="80" text-anchor="middle" font-size="22"
                              font-weight="700" fill="{risk_color}"
                              font-family="JetBrains Mono,monospace">{pct:.1f}%</text>
                        <text x="90" y="100" text-anchor="middle" font-size="10"
                              font-weight="700" fill="{risk_color}"
                              font-family="Plus Jakarta Sans,sans-serif"
                              letter-spacing="2">{risk} RISK</text>
                    </svg>
                    <div style="color:{C_MUTED};font-size:10px;font-weight:600;
                                text-transform:uppercase;letter-spacing:0.1em;margin-top:4px;">
                        FRAUD PROBABILITY</div>
                </div>""", unsafe_allow_html=True)

            with col_d:
                detail_cols = [c for c in [
                    "Transaction_ID","Transaction_Amount","Transaction_Type",
                    "Merchant_Category","Card_Type","Account_Balance",
                    "Is_International","Unusual_Time","Distance_From_Home",
                    "Failed_Transaction_Count","Previous_Fraud_Count",
                ] if c in row.index]
                rows_html = "".join([
                    f'<div style="display:flex;justify-content:space-between;'
                    f'padding:8px 0;border-bottom:1px solid {C_BORDER};font-size:13px;">'
                    f'<span style="color:{C_MUTED};font-weight:600;font-size:11px;'
                    f'text-transform:uppercase;letter-spacing:0.06em;">{c.replace("_"," ")}</span>'
                    f'<span style="color:{C_TEXT};font-family:\'JetBrains Mono\',monospace;'
                    f'font-size:12px;">{row[c]}</span></div>'
                    for c in detail_cols
                ])
                st.markdown(
                    f'<div style="background:{C_BG2};border:1px solid {C_BORDER};'
                    f'border-radius:14px;padding:18px 22px;">{rows_html}</div>',
                    unsafe_allow_html=True)